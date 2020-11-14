from openpyxl import load_workbook
import os
import sys
import pprint
import logging

# path
file_dir = '.\\docs\\'
filename = 'Data.xlsx'
mylogs = 'logs\\MyLogs.log'
file = os.path.join(file_dir, filename)
logging.basicConfig(filename=os.path.join(file_dir,mylogs), level=logging.ERROR, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# STOCK DATA
class LoadStockData:
    def __init__(self, file=file):
        self.wb = load_workbook(file)
        # save all sheets
        self.d1 = self.wb['D1']
        self.d2 = self.wb['D2']
        self.d3 = self.wb['D3']
        self.no_of_trades = self.wb['nooftrades']

# FILTERED STOCK DATA
class MySampleStockFilter:
    def __init__(self):
        self.wb = LoadStockData()
        self.stock_data = {}
        self._get_todays_stock_info()
        self._get_number_of_trades()
        self._get_watchlist()
    
    def _get_todays_stock_info(self):
        # get the info of the current day
        for row in self.wb.d1.iter_rows(min_row=2, min_col=1, max_col=59, values_only=True):
            stock = row[0]
            values = {
                "Last Price": row[1],
                "MA 50": row[18],
                "MA 20": row[17],
                "Value": row[9],
                "ValueSMA(20)": row[41]
            }
            self.stock_data[stock] = values

    def _get_number_of_trades(self):
        # get the number of trades
        for row in self.wb.no_of_trades.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
            num = str(row[4]).strip()
            try:
                if num[-1].lower() == 'k':
                    num = num[:-1]
                    num1 = float(num) * 1000
                else:
                    num1 = float(num)
                self.stock_data[row[0]]["No. of trades"] = round(num1, 2)
            except ValueError:
                logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Getting number of trades",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                pass
            except KeyError:
                logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Getting number of trades",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                pass
    
    def _get_watchlist(self):
        # make a watch list
        self.watchlist = {}
        self.watchlist_no_of_trades = {}

        # get the stocks where Price above MA50 and MA20 above MA50
        for stock in self.stock_data:
            # test for the EOD Reg conditions
            test_cond1 = 0
            try:
                # 1. Price above MA50 and MA20 above MA50
                if (self.stock_data[stock]["Last Price"] > self.stock_data[stock]["MA 50"]) and (self.stock_data[stock]["MA 20"] > self.stock_data[stock]["MA 50"]):
                    test_cond1 += 1
            except KeyError:
                logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Filtering stocks",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                pass
            except TypeError:
                logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Filtering stocks",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                pass
        
        # Determine the volume filter
            try:
                # 1. AND((Value/ValueSMA(20))>=1,Value>=5000000,ValueSMA(20)<5000000),"Volume Break-out"
                if (self.stock_data[stock]["Value"]/self.stock_data[stock]["ValueSMA(20)"]>=1) and (self.stock_data[stock]["Value"]>=5000000) and (self.stock_data[stock]["ValueSMA(20)"]<5000000):
                    volume_filter = "Volume Break-out"
                # 2. AND((Value/ValueSMA(20))>=1,Value>=5000000,ValueSMA(20)>=5000000),"VolBreak-out+"
                elif (self.stock_data[stock]["Value"]/self.stock_data[stock]["ValueSMA(20)"]>=1) and (self.stock_data[stock]["Value"]>=5000000) and (self.stock_data[stock]["ValueSMA(20)"]>=5000000):
                    volume_filter = "VolBreak-out+"
                # 3. ValueSMA(20)>=5000000,">5MVolVal20SMA"
                elif self.stock_data[stock]["ValueSMA(20)"] >= 5000000:
                    volume_filter = ">5MVolVal20SMA"
                # 4. Value>=20000000,">20MDailyVol"
                elif self.stock_data[stock]["Value"] >= 20000000:
                    volume_filter = ">20MDailyVol"
                else:
                    volume_filter = ""
            except ZeroDivisionError:
                logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Determining volume filter",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                pass
            except TypeError:
                logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Determining volume filter",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                pass
        
        # get the stocks that satisfie the entry conditions
            if test_cond1 == 1:
                try:
                    cond1_data = {
                        "Strategy": 'Trend Following',
                        "No. of Trades": self.stock_data[stock]["No. of trades"],
                        "Volume Filter": volume_filter
                    }
                except KeyError:
                    logging.error("State: {0:^35s} <====> Error Class:{1:^20s} <====> Details: {2:<20s}".format("Getting Filtered Stocks",str(sys.exc_info()[0]), str(sys.exc_info()[1])))
                    pass
                else:
                    self.watchlist[stock] = cond1_data
                    self.watchlist_no_of_trades[stock] = self.stock_data[stock]["No. of trades"]

